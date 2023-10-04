import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import datetime
import time
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# current_day_name = datetime.datetime.now().strftime('%A')
current_day_name = 'Friday'

excel_file = "Excel.xlsx"
workbook = openpyxl.load_workbook(excel_file)

worksheet = workbook[current_day_name]

df = pd.DataFrame(worksheet.values)
print(df)

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

driver.get('https://www.google.com')
driver.maximize_window()
search_box = driver.find_element(By.NAME, 'q')

for row in range(len(df.values)):  # This creates a range from 2 to 12
    keyword = df.at[row, 2]
    search_box.clear()
    if keyword is not None:
        search_box.send_keys(keyword)
        time.sleep(2)
        suggestions = driver.find_elements(By.TAG_NAME, 'li')
        suggestions = suggestions[0:len(suggestions)-2]
        shortest_suggestion = min(suggestions, key=lambda x: len(x.text)).text
        longest_suggestion = max(suggestions, key=lambda x: len(x.text)).text

        print('Shortest Suggestion:', shortest_suggestion)
        print('Longest Suggestion:', longest_suggestion)

        worksheet.cell(row=row+1, column=4, value=longest_suggestion)
        worksheet.cell(row=row+1, column=5, value=shortest_suggestion)

workbook.save(excel_file)
driver.quit()







# import openpyxl
# import pandas as pd
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# import datetime
# import time
# from selenium.webdriver.chrome.service import Service as ChromeService
# from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

# # Get the current day name
# current_day_name = datetime.datetime.now().strftime('%A')

# # Load the Excel file
# excel_file = 'Excel.xlsx'
# workbook = openpyxl.load_workbook(excel_file)

# # Get the worksheet for the current day
# worksheet = workbook[current_day_name]
# print(workbook)

# # Read the worksheet into a pandas DataFrame
# df = pd.DataFrame(worksheet.values)

# # Get the keyword from the DataFrame
# keyword = df.at[4, 2]  # Assuming the keyword is in the first cell

# # Initialize the WebDriver (make sure to set the path to your ChromeDriver executable)
# driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

# # Open Google and search for the keyword
# driver.get('https://www.google.com')
# driver.maximize_window()
# search_box = driver.find_element(By.NAME, 'q')
# search_box.clear()
# search_box.send_keys(keyword)
# time.sleep(3)

# x_path = "/html/body/div[2]/div[2]/div[2]/div[2]/form/div[1]/div[1]/div[2]/div[1]/div/ul/li[4]/div[1]/div[2]/div[1]/span"
# suggestions = driver.find_elements(By.TAG_NAME, 'li')
# print(suggestions)


# # Wait for search suggestions to appear
# print(len(suggestions))

# # Capture and print the suggestions
# for suggestion in suggestions:
#     suggestion_text = suggestion.text
#     print(f'Suggestion: {suggestion_text}')

# # Find the shortest and longest suggestion
# shortest_suggestion = min(suggestions, key=lambda x: len(x.text)).text
# longest_suggestion = max(suggestions, key=lambda x: len(x.text)).text

# print('Shortest Suggestion:', shortest_suggestion)
# print('Longest Suggestion:', longest_suggestion)

# # Update the DataFrame with the shortest and longest suggestions
# df.at[3, 3] = shortest_suggestion
# df.at[3, 4] = longest_suggestion

# # Save the updated DataFrame back to the worksheet
# for r_idx, row in enumerate(df.values):
#     for c_idx, value in enumerate(row):
#         worksheet.cell(row=r_idx + 1, column=c_idx + 1, value=value)

# # Save the Excel file
# workbook.save(excel_file)
# print(worksheet.values)

# # Close the WebDriver
# driver.quit()

